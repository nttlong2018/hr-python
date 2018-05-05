from openpyxl import Workbook
import openpyxl
import bson
def write_to(file_name,data,fields):
    wb = Workbook()
    ws_mapping_sheet = wb.create_sheet("mapping")
    index=1
    for item in fields:
        ws_mapping_sheet.cell(index,1,item["field"])
        ws_mapping_sheet.cell(index, 2, item["type"])
        index+=1
    ws_main=wb.create_sheet("main")
    for row in data:
        data_row=[]
        for key in row.keys():
            if type(row[key]) is bson.objectid.ObjectId:
                data_row.append(row[key].__str__())
            else:
                data_row.append(row[key])
        ws_main.append(data_row)

    wb.save(file_name)
def get_coll_address(str):
    ret=""
    for x in str:
        if "1234567890".find(x)==-1:
            ret+=x
        else:
            return ret
def read_from_file(file_name):
    wb =openpyxl.load_workbook(file_name)
    ws_mapping=wb.get_sheet_by_name("mapping")
    if ws_mapping==None:
        wb.close()
        raise "mapping sheet was not found in '{0}'".format(file_name)
    _model={}
    # cols=list(ws_mapping.columns)
    cells=list(wb.defined_names.definedName)
    hash_columns={}
    for row in cells:
        field_name=row.name
        items=field_name.split('.')
        field=_model
        for item in items:
            if not field.has_key(item):
                if items.index(item) < items.__len__()-1:
                    field.update({item:{}})
                    field=field[item]
                else:
                    # info=dict(
                    #     sheet=row[2].split('=')[1].split('!')[0],
                    #     address=row[2].split('=')[1].split('!')[1],
                    #     col_address=get_coll_address(row[2].split('=')[1].split('!')[1])
                    # )
                    hash_columns.update({
                        field_name:{
                           "address": row.value.split("$")[1],
                            "fields":items
                        }

                    })
                    field.update({item: ""})
            else:
                field = field[item]

        field=""
    ws_main = wb.get_sheet_by_name("main")
    row_count=ws_main.max_row
    ret=[]

    for i in range(2,row_count+1):
        data_row=_model.copy()
        for key in hash_columns.keys():
            keys=hash_columns[key]["fields"]
            value=data_row
            for x in keys:
                if keys.index(x)<keys.__len__()-1:
                    value=value[x]
                else:
                    val=ws_main[hash_columns[key]["address"]+i.__str__()].value
                    value.update({
                        x:val
                    })
        ret.append(data_row)



    return dict(
        data=ret,
        model=_model
    )




    # for row in range(2, ws_mapping.max_row + 1):
    #     pass






