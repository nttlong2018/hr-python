from django.http import HttpResponse
import quicky
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
import datetime
import base64
from io import BytesIO
import constant as KEY


def call(args):
    wb = load_workbook(filename=BytesIO(base64.b64decode(args["data"]["_content"])))
    #check exists mapping sheet
    if KEY.WS_MAPPING_SHEET in wb.sheetnames:
        ws_mapping = wb[KEY.WS_MAPPING_SHEET]
        mapping_configs = __get_mapping(ws_mapping)
        data = []
        for map_item in mapping_configs:
            ws_data = wb[map_item['collection_name']]
            _data = __get_data(mapping_configs, ws_data)
            data.append(_data)

        return data
    else:
        return {
            error: "File structure error!"
        }

def __get_data(mapping_configs, ws_data):
    return {'data': None}

def __get_mapping(ws_mapping):
    "Read data in mapping sheet and convert to python object"
    mapping_configs = []
    config = {}
    for idx_row in range(1, ws_mapping.max_row + 1):
        cell_col1 = ws_mapping.cell(row=idx_row, column=1)
        cell_col2 = ws_mapping.cell(row=idx_row, column=2)
        if(cell_col1.value == KEY.BEGIN_MAPPING):
            config = {
                'collection_name' : ws_mapping.cell(row=idx_row, column=3).value,
                'fields' : {}
            }
        elif (cell_col1.value == KEY.END_MAPPING):
            mapping_configs.append(config)
            config = {}
        else:
            config['fields'][cell_col1.value] = cell_col2.value
    return mapping_configs