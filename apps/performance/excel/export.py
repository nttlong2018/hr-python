# -*- coding: utf-8 -*-
from django.http import HttpResponse
import quicky
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import datetime
import performance.api.models as models
from format_worksheet import worksheet_style, format_style 
from openpyxl.workbook import defined_name

@quicky.view.template("")
def call(request):
    wb = Workbook()
    ws = wb.active
    #ws["A1"].value = "TEST"
    ws.title = "HCSSYS_DataDomain";

    #Get column config 
    coll = models.HCSSYS_CollectionInfo().aggregate()
    coll._pipe.append({
        '$match': { 
            '$and' : [
                {'parent_field' : None}, 
                {'field_path': {'$regex': '^,HCSSYS_DataDomain,'}},
                {'is_parent' : False}, 
               ] 
            }
        });
    coll._pipe.append({'$project': {'_id': 0}})
    columns = coll.get_list()

    #Get export data
    data_items = [];
    try:
        __collection = getattr(models, "HCSSYS_DataDomain")()
        data_items = __collection .get_list()
    except Exception as ex:
        raise ex

    #Create header worksheet (row 1)
    for iCol, vCol in enumerate(columns):
        cell = ws.cell(row=1, column=iCol + 1)
        cell.value = vCol["description"]
        #cell_address = ws.title + "!$" + cell.column + "$" + str(cell.row)
        cell_address = ws.title + "!$" + cell.column + "$1"
        wb.defined_names.append(defined_name.DefinedName(attr_text=cell_address, name=vCol["field_name"]))


    #Render content to worksheet
    #if (len(data_items) > 0):
    for iItem, vItem in enumerate(data_items):
        data_row = [];
        for iCol, vCol in enumerate(columns):
            data_row.append(vItem[vCol["field_name"]])
        ws.append(data_row) 

    #format worksheet
    ws = format_style(worksheet_style["NORMAL"]).format(ws)

    wb.defined_names.append(defined_name.DefinedName(attr_text="HCSSYS_DataDomain!A$1", name="TEST_NAME_0000000"))


    response = HttpResponse(content=save_virtual_workbook(wb), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + 'HCSSYS_DataDomain' + '.xlsx'
    return response
